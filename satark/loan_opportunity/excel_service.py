import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_call_analysis_excel(rows_data):
    """
    Generates an Excel workbook (.xlsx) binary stream for Loan Disbursal & EMI Collection Analysis.
    
    Columns:
      1. Call Date
      2. Client / Customer Number
      3. Agent / Staff Number
      4. Customer Info ID
      5. Disbursement ID
      6. User ID
      7. Branch ID
      8. Circle / Operator
      9. Customer Ready to Pay (1/0)
      10. New / Top-Up Loan Interest (1/0)
      11. Promised EMI Amount
      12. Promised PTP Date
      13. Reason for Non-Payment / Delay
      14. Customer Financial Situation
      15. Collection Outcome
      16. Recommended BRO / Manager Action
      17. Raw Hindi Transcript
      18. Staff & User Interaction
      19. English Executive Summary
      20. Recording Audio URL
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Opportunity Analysis"

    headers = [
        "Call Date", 
        "Client / Customer Number", 
        "Agent / Staff Number", 
        "Customer Info ID", 
        "Disbursement ID", 
        "User ID", 
        "Branch ID", 
        "Circle / Operator", 
        "Customer Ready to Pay (1/0)",
        "New / Top-Up Loan Interest (1/0)",
        "Customer Referral Interest (1/0)",
        "Referred Customer Details",
        "STT Transcript Accuracy (%)",
        "Ready to Pay Confidence (%)",
        "New Loan Confidence (%)",
        "Referral Confidence (%)",
        "Overall Call Confidence (%)",
        "Confidence Grade",
        "Promised EMI Amount", 
        "Promised PTP Date", 
        "Reason for Non-Payment / Delay", 
        "Customer Financial Situation", 
        "Collection Outcome", 
        "Recommended BRO / Manager Action", 
        "Raw Hindi Transcript", 
        "Staff & User Interaction", 
        "English Executive Summary", 
        "Recording Audio URL",
        "Is Fallback Used? (1/0)",
        "Fallback Reason / Audit Notes"
    ]

    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # Status fills
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid") # Light Green
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")   # Light Red
    purple_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid") # Light Purple for New Loan Interest
    amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber for Referral Interest/Fallback

    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Add Data Rows
    for row_idx, r in enumerate(rows_data, start=2):
        ready_val = r.get("ready_to_pay", 1)
        new_loan_val = r.get("new_loan_interest", 0)
        referral_val = r.get("referral_interest", 0)
        fallback_val = r.get("is_fallback", 0)

        stt_conf = r.get("stt_transcript_confidence", 80)
        rtp_conf = r.get("ready_to_pay_confidence", 85)
        new_loan_conf = r.get("new_loan_confidence", 60)
        ref_conf = r.get("referral_confidence", 95)
        overall_conf = r.get("overall_call_confidence", 75)
        conf_grade = r.get("confidence_grade", "MEDIUM")

        row_values = [
            r.get("date", ""),
            r.get("client_number", ""),
            r.get("agent_number", ""),
            r.get("customerinfoid", ""),
            r.get("disbursementid", ""),
            r.get("userid", ""),
            r.get("branchhid", ""),
            f"{r.get('circle_operator', '')} - {r.get('circle_circle', '')}",
            ready_val,
            new_loan_val,
            referral_val,
            r.get("referred_customer_details", ""),
            f"{stt_conf}%",
            f"{rtp_conf}%",
            f"{new_loan_conf}%",
            f"{ref_conf}%",
            f"{overall_conf}%",
            conf_grade,
            r.get("promised_amount", ""),
            r.get("promised_date", ""),
            r.get("reason_for_non_payment", ""),
            r.get("customer_situation", ""),
            r.get("collection_outcome", ""),
            r.get("recommended_bro_action", ""),
            r.get("raw_transcript", ""),
            r.get("staff_user_interaction", ""),
            r.get("english_summary", ""),
            r.get("recording_url", ""),
            fallback_val,
            r.get("fallback_reason", "")
        ]
        ws.append(row_values)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = cell_font
            cell.border = thin_border
            
            # Alignments
            if col_num in [1, 2, 3, 4, 5, 6, 7, 9, 10, 11, 13, 14, 15, 16, 17, 18, 20, 29]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # Highlight Ready to Pay column (Col 9)
            if col_num == 9:
                if ready_val == 1:
                    cell.fill = green_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="065F46")
                else:
                    cell.fill = red_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B")

            # Highlight New / Top-Up Loan Interest column (Col 10)
            if col_num == 10:
                if new_loan_val == 1:
                    cell.fill = purple_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="6B21A8")
                else:
                    cell.font = Font(name="Calibri", size=10, color="6B7280")

            # Highlight Customer Referral Interest column (Col 11)
            if col_num == 11:
                if referral_val == 1:
                    cell.fill = amber_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")
                else:
                    cell.font = Font(name="Calibri", size=10, color="6B7280")

            # Highlight Confidence Grade column (Col 18)
            if col_num == 18:
                if conf_grade == "HIGH":
                    cell.fill = green_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="065F46")
                elif conf_grade == "MEDIUM":
                    cell.fill = amber_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")
                else:
                    cell.fill = red_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="991B1B")

            # Highlight Is Fallback Used column (Col 29)
            if col_num == 29:
                if fallback_val == 1:
                    cell.fill = amber_fill
                    cell.font = Font(name="Calibri", size=10, bold=True, color="B45309")
                else:
                    cell.font = Font(name="Calibri", size=10, color="6B7280")

    # Column Widths
    col_widths = {
        1: 14, 2: 18, 3: 18, 4: 16, 5: 16, 6: 12, 7: 12, 8: 24,
        9: 16, 10: 20, 11: 20, 12: 25, 13: 18, 14: 18, 15: 18, 16: 18,
        17: 20, 18: 16, 19: 18, 20: 16, 21: 30, 22: 35, 23: 20, 24: 35,
        25: 45, 26: 45, 27: 45, 28: 30, 29: 18, 30: 35
    }
    
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
