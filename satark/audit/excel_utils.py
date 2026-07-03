import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection


def _protect_sheet(ws):
    """
    Lock the worksheet so it is view-only.
    - Sheet is protected (no edits, no inserts, no deletes, no formatting).
    - AutoFilter (column filters) is the ONLY operation permitted.
    - No password is required to VIEW; the protection is passwordless but
      enforced by Excel — users simply cannot edit any cell.
    """
    ws.protection = SheetProtection(
        sheet=True,           # enable protection
        password='',          # no password needed to turn it off, but sheet is still locked
        selectLockedCells=True,
        selectUnlockedCells=True,
        autoFilter=False,     # False here means "do NOT restrict autoFilter" → filtering is allowed
        sort=True,            # True means "restrict sort" → sorting blocked (change to False to allow)
        insertRows=True,
        insertColumns=True,
        deleteRows=True,
        deleteColumns=True,
        formatCells=True,
        formatColumns=True,
        formatRows=True,
        pivotTables=True,
    )


def generate_branch_audit_excel(metadata, branch_points, center_points, client_points):
    wb = openpyxl.Workbook()
    
    # 1. Branch Sheet
    ws_branch = wb.active
    ws_branch.title = "Branch"
    
    # Define styles
    header_fill = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border_style = Side(border_style="thin", color="000000")
    thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    branch_headers = [
        "Section", "Intent", "Key Risk Issues", "Category", 
        "Yes / No / NA", "Sample Verification", "Correct Finding", "Wrong Finding",
        "Total Sample", "Process Deviation %", "Max Score", "Obtained Score",
        "Is Issue", "Auditor Remark", "Reviewer Remark"
    ]
    
    ws_branch.append(branch_headers)
    for col_num, header in enumerate(branch_headers, 1):
        cell = ws_branch.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = alignment_center

    for point in branch_points:
        row = [
            point.get('section_code', ''),
            point.get('intent', ''),
            point.get('risk_issue', ''),
            point.get('category', ''),
            point.get('yes_no_na', ''),
            point.get('sample_verification', ''),
            point.get('correct_finding', ''),
            point.get('wrong_finding', ''),
            point.get('total_sample', ''),
            point.get('process_deviation', ''),
            point.get('max_score', ''),
            point.get('obtained_score', ''),
            point.get('is_issue', ''),
            point.get('auditor_remark', '') or '',
            point.get('reviewer_remark', '') or ''
        ]
        ws_branch.append(row)
        for col_num in range(1, len(row) + 1):
            cell = ws_branch.cell(row=ws_branch.max_row, column=col_num)
            cell.border = thin_border
            cell.alignment = alignment_left

    # Adjust column widths for Branch
    for col in ws_branch.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 5, 50) # Cap at 50
        ws_branch.column_dimensions[column].width = adjusted_width

    # Add AutoFilter to Branch header row and protect the sheet
    ws_branch.auto_filter.ref = ws_branch.dimensions
    _protect_sheet(ws_branch)


    # 2. Center Sheet
    ws_center = wb.create_sheet(title="Center")
    
    # We want Centers as Rows and Parameters as Columns
    # First, get all unique parameters for centers
    center_params = {}
    for pt in center_points:
        p_code = pt.get('parameter_code')
        if p_code not in center_params:
            center_params[p_code] = pt.get('parameter_name', p_code)
            
    center_param_codes = list(center_params.keys())
    
    center_headers = ["Center ID"] + [center_params[p] for p in center_param_codes] + ["Auditor Remark", "Reviewer Remark"]
    ws_center.append(center_headers)
    
    for col_num, header in enumerate(center_headers, 1):
        cell = ws_center.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = alignment_center

    # Group by center
    centers = {}
    for pt in center_points:
        cid = pt.get('center_id')
        if cid not in centers:
            centers[cid] = {'answers': {}, 'auditor_remark': '', 'reviewer_remark': ''}
        centers[cid]['answers'][pt.get('parameter_code')] = pt.get('answer', '')
        if pt.get('auditor_remark'):
            centers[cid]['auditor_remark'] = pt.get('auditor_remark', '') or ''
        if pt.get('reviewer_remark'):
            centers[cid]['reviewer_remark'] = pt.get('reviewer_remark', '') or ''
        
    for cid, data in centers.items():
        row = [str(cid)]
        for p_code in center_param_codes:
            row.append(data['answers'].get(p_code, ''))
        row.append(data['auditor_remark'])
        row.append(data['reviewer_remark'])
        ws_center.append(row)
        for col_num in range(1, len(row) + 1):
            cell = ws_center.cell(row=ws_center.max_row, column=col_num)
            cell.border = thin_border
            cell.alignment = alignment_center

    for col in ws_center.columns:
        ws_center.column_dimensions[col[0].column_letter].width = 25

    # Add AutoFilter to Center header row and protect the sheet
    ws_center.auto_filter.ref = ws_center.dimensions
    _protect_sheet(ws_center)


    # 3. Client Sheet
    ws_client = wb.create_sheet(title="Client")
    
    client_params = {}
    for pt in client_points:
        p_code = pt.get('parameter_code')
        if p_code not in client_params:
            client_params[p_code] = pt.get('parameter_name', p_code)
            
    client_param_codes = list(client_params.keys())
    
    client_headers = ["Center ID", "Client ID", "Client Name"] + [client_params[p] for p in client_param_codes] + ["Auditor Remark", "Reviewer Remark"]
    ws_client.append(client_headers)
    
    for col_num, header in enumerate(client_headers, 1):
        cell = ws_client.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = alignment_center

    # Group by client
    clients = {}
    for pt in client_points:
        client_key = (pt.get('center_id'), pt.get('client_id'), pt.get('client_name'))
        if client_key not in clients:
            clients[client_key] = {'answers': {}, 'auditor_remark': '', 'reviewer_remark': ''}
        clients[client_key]['answers'][pt.get('parameter_code')] = pt.get('answer', '')
        if pt.get('auditor_remark'):
            clients[client_key]['auditor_remark'] = pt.get('auditor_remark', '') or ''
        if pt.get('reviewer_remark'):
            clients[client_key]['reviewer_remark'] = pt.get('reviewer_remark', '') or ''
        
    for c_key, data in clients.items():
        row = [str(c_key[0]), str(c_key[1]), str(c_key[2])]
        for p_code in client_param_codes:
            row.append(data['answers'].get(p_code, ''))
        row.append(data['auditor_remark'])
        row.append(data['reviewer_remark'])
        ws_client.append(row)
        for col_num in range(1, len(row) + 1):
            cell = ws_client.cell(row=ws_client.max_row, column=col_num)
            cell.border = thin_border
            cell.alignment = alignment_center

    for col in ws_client.columns:
        ws_client.column_dimensions[col[0].column_letter].width = 25

    # Add AutoFilter to Client header row and protect the sheet
    ws_client.auto_filter.ref = ws_client.dimensions
    _protect_sheet(ws_client)


    # 4. Final Combined Result
    ws_final = wb.create_sheet(title="Final Combined Result")
    
    ws_final.append(["Final Combined Result"])
    title_cell = ws_final.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=16)
    
    ws_final.append(["Branch Name", metadata.get('branch_name', '')])
    ws_final.append(["Audit Period", metadata.get('audit_period', '')])
    ws_final.append(["Auditor", metadata.get('auditor_name', '')])
    ws_final.append([])
    
    ws_final.append(["Scores"])
    ws_final.cell(row=ws_final.max_row, column=1).font = Font(bold=True, size=14)
    ws_final.append(["Total Score", metadata.get('total_score', '')])
    ws_final.append(["Total Max Score", metadata.get('total_max_score', '')])
    ws_final.append(["Percentage", f"{metadata.get('percentage', '')}%"])
    ws_final.append(["Grade", metadata.get('grade', '')])

    # Protect the Final Combined Result sheet (view-only, no filters needed here)
    _protect_sheet(ws_final)

    return wb
