import json
import datetime
import logging
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import jwt

from satark.cqrs import dispatcher
from .models import AccountsMstUsertbl, JWTToken
from .utils import log_info, log_error
from .commands import (
    LoginCommand, LoginCommandHandler,
    RefreshCommand, RefreshCommandHandler,
    SaveMenuCommand, SaveMenuCommandHandler,
    SaveUserCommand, SaveUserCommandHandler,
    CreateRoleCommand, CreateRoleCommandHandler,
    MapUserRoleCommand, MapUserRoleCommandHandler,
    CreateUserCommand, CreateUserCommandHandler,
)
from .queries import (
    GetMenuQuery, GetMenuQueryHandler,
    GetAdminMenuQuery, GetAdminMenuQueryHandler,
    GetUserListQuery, GetUserListQueryHandler,
    GetGeoHierarchyQuery, GetGeoHierarchyQueryHandler
)

logger = logging.getLogger("authentication.views")

# Register commands and queries with the dispatcher
dispatcher.register_command(LoginCommand, LoginCommandHandler())
dispatcher.register_command(RefreshCommand, RefreshCommandHandler())
dispatcher.register_command(SaveMenuCommand, SaveMenuCommandHandler())
dispatcher.register_command(SaveUserCommand, SaveUserCommandHandler())
dispatcher.register_command(CreateRoleCommand, CreateRoleCommandHandler())
dispatcher.register_command(MapUserRoleCommand, MapUserRoleCommandHandler())
dispatcher.register_command(CreateUserCommand, CreateUserCommandHandler())

dispatcher.register_query(GetMenuQuery, GetMenuQueryHandler())
dispatcher.register_query(GetAdminMenuQuery, GetAdminMenuQueryHandler())
dispatcher.register_query(GetUserListQuery, GetUserListQueryHandler())
dispatcher.register_query(GetGeoHierarchyQuery, GetGeoHierarchyQueryHandler())


# --- Reusable Security Helpers ---

def validate_token_user(token):
    try:
        jwt_payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
        user_id = jwt_payload.get('user_id')
        return AccountsMstUsertbl.objects.get(id=user_id)
    except Exception:
        return None

def is_user_admin(user):
    from django.db import connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM [dbo].[map_userRole] WHERE UserID = %s AND RoleId = 1 AND IsActive = 1", [user.UserID])
            return cursor.fetchone()[0] > 0
    except Exception:
        return False

def send_encrypted_response(data_dict, status_code=200):
    return JsonResponse(data_dict, status=status_code)


# --- View Endpoints ---

@csrf_exempt
def login_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body) if request.body else {}
        usercode = data.get('username', '').strip()
        password = data.get('password', '')
    except Exception as e:
        log_error(f"Login request parsing failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    if not usercode or not password:
        log_error("Login failed: usercode or password missing")
        return send_encrypted_response({'success': False, 'message': 'usercode and password are required'}, status_code=400)

    try:
        command = LoginCommand(usercode=usercode, password=password)
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Login view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def refresh_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        refresh_token = data.get('refresh_token', '')
    except Exception as e:
        log_error(f"Refresh request parsing failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    if not refresh_token:
        log_error("Refresh failed: Refresh token missing")
        return send_encrypted_response({'success': False, 'message': 'Refresh token is required'}, status_code=400)

    try:
        command = RefreshCommand(refresh_token=refresh_token)
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Refresh view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def menu_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        log_error(f"Menu request parsing failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)

    try:
        query = GetMenuQuery(user_db_id=user.UserID)
        result = dispatcher.query(query)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Menu view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def admin_menu_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        log_error(f"Admin menu request parsing failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)

    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        query = GetAdminMenuQuery()
        result = dispatcher.query(query)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Admin menu view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def admin_save_menu_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
        items = data.get('items', [])
        mappings = data.get('mappings', [])
    except Exception as e:
        log_error(f"Admin menu save request parsing failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)

    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        command = SaveMenuCommand(items=items, mappings=mappings)
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Admin menu save view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def admin_user_list_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        log_error(f"admin_user_list_view: parse failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)

    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        query = GetUserListQuery()
        result = dispatcher.query(query)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Admin user list view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def admin_geo_hierarchy_view(request):
    """Lazy-loaded geographic hierarchy — called only when edit modal opens"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        log_error(f"admin_geo_hierarchy_view: parse failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)

    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        query = GetGeoHierarchyQuery()
        result = dispatcher.query(query)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Admin geo hierarchy view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def admin_save_user_view(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
        user_id = data.get('user_id')
        new_role_id = data.get('role_id')
        new_branch_id = data.get('branch_id')
        new_contact = data.get('contact_no')
        new_email = data.get('email')
    except Exception as e:
        log_error(f"admin_save_user_view: parse failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    admin_user = validate_token_user(token)
    if not admin_user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)

    if not is_user_admin(admin_user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    if not user_id:
        return send_encrypted_response({'success': False, 'message': 'user_id is required'}, status_code=400)

    try:
        command = SaveUserCommand(
            user_id=user_id,
            role_id=new_role_id,
            branch_id=new_branch_id,
            contact_no=new_contact,
            email=new_email
        )
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"Admin save user view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def create_role_view(request):
    """Create a new role in mst_role"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)
    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        command = CreateRoleCommand(
            role_name=data.get('role_name', ''),
            description=data.get('description', '')
        )
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"create_role_view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def map_user_role_view(request):
    """Map a UserID to a RoleID in map_userrole"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)
    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        command = MapUserRoleCommand(
            user_id=data.get('user_id'),
            role_id=data.get('role_id'),
            is_active=data.get('is_active', True)
        )
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"map_user_role_view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


@csrf_exempt
def create_user_view(request):
    """Create a new user in accounts_mst_usertbl with role mapping"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        token = data.get('token', '')
    except Exception as e:
        return send_encrypted_response({'success': False, 'message': 'Invalid request body'}, status_code=400)

    user = validate_token_user(token)
    if not user:
        return send_encrypted_response({'success': False, 'message': 'Invalid token'}, status_code=401)
    if not is_user_admin(user):
        return send_encrypted_response({'success': False, 'message': 'Access denied'}, status_code=403)

    try:
        command = CreateUserCommand(
            user_id=data.get('user_id'),
            user_name=data.get('user_name', ''),
            user_code=data.get('user_code', ''),
            contact_no=data.get('contact_no', ''),
            email=data.get('email', ''),
            branch_id=data.get('branch_id'),
            role_id=data.get('role_id'),
            is_active=data.get('is_active', True)
        )
        result = dispatcher.send(command)
        status_code = result.pop('status_code', 200)
        return send_encrypted_response(result, status_code=status_code)
    except Exception as e:
        log_error(f"create_user_view failed: {str(e)}")
        return send_encrypted_response({'success': False, 'message': 'Internal Server Error'}, status_code=500)


# --- USER TRACKING & ANALYTICS VIEWS ---

from .user_tracking import (
    record_session_start, record_session_heartbeat, record_session_end,
    record_screen_log, get_analytics_overview, get_screen_analytics,
    get_api_analytics, get_user_analytics_list, get_live_users, get_session_logs
)
import csv
from django.http import HttpResponse

@csrf_exempt
def session_start_view(request):
    """Initializes tracking session upon login or page refresh"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        user_id = data.get('user_id')
        usercode = data.get('usercode', '')
        if not user_id or not usercode:
            return JsonResponse({'success': False, 'message': 'Missing user parameters'}, status=400)
        
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        ip_addr = request.META.get('REMOTE_ADDR', '')
        session_id = record_session_start(user_id, usercode, ip_addr, user_agent)
        return JsonResponse({'success': True, 'session_id': session_id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def session_heartbeat_view(request):
    """Heartbeat endpoint called by frontend every 30 seconds"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        session_id = data.get('session_id')
        active_delta = data.get('active_delta', 0)
        idle_delta = data.get('idle_delta', 0)
        
        if session_id:
            record_session_heartbeat(session_id, active_delta, idle_delta)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def session_end_view(request):
    """End session endpoint (manual logout or 15-min idle timeout)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        session_id = data.get('session_id')
        reason = data.get('reason', 'MANUAL')
        if session_id:
            record_session_end(session_id, reason)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def session_screen_log_view(request):
    """Log screen transition & active time on frontend screen"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body) if request.body else {}
        session_id = data.get('session_id')
        user_id = data.get('user_id')
        usercode = data.get('usercode', '')
        screen_name = data.get('screen_name', '')
        path = data.get('path', '')
        active_sec = data.get('active_sec', 0)
        idle_sec = data.get('idle_sec', 0)

        record_screen_log(session_id, user_id, usercode, screen_name, path, active_sec, idle_sec)
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def admin_analytics_dashboard_view(request):
    """Returns high-level KPI overview & summary telemetry for admin user"""
    try:
        overview = get_analytics_overview()
        session_logs = get_session_logs(limit=25)
        return JsonResponse({'success': True, 'overview': overview, 'recent_sessions': session_logs})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def admin_analytics_screens_view(request):
    """Returns screen time & visit analytics"""
    try:
        screens = get_screen_analytics()
        return JsonResponse({'success': True, 'screens': screens})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def admin_analytics_apis_view(request):
    """Returns most frequently called APIs telemetry"""
    try:
        apis = get_api_analytics()
        return JsonResponse({'success': True, 'apis': apis})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def admin_analytics_users_view(request):
    """Returns user active duration & session breakdown"""
    try:
        users = get_user_analytics_list()
        return JsonResponse({'success': True, 'users': users})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
def admin_analytics_live_view(request):
    """Returns currently online active users"""
    try:
        live = get_live_users()
        return JsonResponse({'success': True, 'live_users': live})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def _style_excel_worksheet(ws, report_title, headers, data_rows, now_str):
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    subtitle_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    odd_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    accent_amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=9, italic=True, color="CBD5E1")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10, color="0F172A")
    font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")

    thin_border_side = Side(border_style="thin", color="E2E8F0")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    num_cols = max(len(headers), 1)
    last_col_letter = get_column_letter(num_cols)

    # 1. Main Title Banner
    ws.merge_cells(f"A1:{last_col_letter}1")
    t_cell = ws["A1"]
    t_cell.value = report_title
    t_cell.font = font_title
    t_cell.fill = title_fill
    t_cell.alignment = align_center
    ws.row_dimensions[1].height = 32

    # 2. Subtitle Meta Banner
    ws.merge_cells(f"A2:{last_col_letter}2")
    s_cell = ws["A2"]
    s_cell.value = f"Generated On: {now_str} IST  |  PRAHARI Telemetry Engine  |  Confidential Enterprise Audit"
    s_cell.font = font_subtitle
    s_cell.fill = subtitle_fill
    s_cell.alignment = align_center
    ws.row_dimensions[2].height = 20

    # 3. Blank Spacing Row
    ws.row_dimensions[3].height = 10

    # 4. Table Headers
    ws.row_dimensions[4].height = 26
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = h_text
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border

    # 5. Populate Data Rows
    for row_idx, r_data in enumerate(data_rows, start=5):
        ws.row_dimensions[row_idx].height = 22
        c_fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
        for col_idx, val in enumerate(r_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = font_data
            cell.fill = c_fill
            cell.border = cell_border

            if isinstance(val, (int, float)):
                cell.alignment = align_right
            elif str(val).startswith('http') or '/' in str(val):
                cell.alignment = align_left
            else:
                cell.alignment = align_center

            if str(val) in ['IDLE_TIMEOUT', '404', '500']:
                cell.fill = accent_amber_fill
                cell.font = font_bold

    # 6. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)


@csrf_exempt
def admin_analytics_export_view(request):
    """Generates professionally styled Excel (.xlsx) telemetry reports including multi-sheet Master reports"""
    report_type = request.GET.get('type', 'master')
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.Workbook()

    if report_type in ['master', 'all']:
        # 1. Sheet 1: Overview & Key Performance Indicators
        ws_ov = wb.active
        ws_ov.title = "Overview & KPIs"
        overview = get_analytics_overview()
        ov_headers = ['Metric Category', 'Value / Summary', 'Description']
        ov_rows = [
            ['Total Sessions Logged', overview.get('total_sessions', 0), 'Total user login sessions recorded in system'],
            ['Currently Active Users', overview.get('currently_active_users', 0), 'Users with heartbeat activity within last 2 minutes'],
            ['Total Active Duration (Hours)', overview.get('total_active_hours', 0.0), 'Combined user active duration in hours'],
            ['Average Session Duration (Mins)', overview.get('avg_session_duration_mins', 0.0), 'Average active duration per user session'],
            ['Top Visited Screen', overview.get('top_screen', 'N/A'), 'Most frequently viewed application screen'],
            ['Top API Endpoint', overview.get('top_api', 'N/A'), 'Most called business API endpoint'],
            ['Total Screen Views', overview.get('total_screen_views', 0), 'Total page/screen navigation view count'],
            ['Total Business API Calls', overview.get('total_api_calls', 0), 'Total business API requests processed'],
        ]
        _style_excel_worksheet(ws_ov, "SATARK AUDIT - MASTER ANALYTICS OVERVIEW & KPIS", ov_headers, ov_rows, now_str)

        # 2. Sheet 2: User Telemetry
        ws_u = wb.create_sheet(title="User Telemetry")
        u_headers = ['User ID', 'UserCode', 'Total Sessions', 'Total Active Mins', 'Last Login', 'Last Active', 'Top Visited Screen']
        u_rows = [[u['user_id'], u['usercode'], u['total_sessions'], u['total_active_mins'], u['last_login'], u['last_active'], u['top_screen']] for u in get_user_analytics_list()]
        _style_excel_worksheet(ws_u, "SATARK AUDIT - USER TELEMETRY & ACTIVE TIME", u_headers, u_rows, now_str)

        # 3. Sheet 3: Screen Usage
        ws_s = wb.create_sheet(title="Screen Usage")
        s_headers = ['Screen Name', 'Route Path', 'Visit Count', 'Unique Users', 'Total Active Mins', 'Avg Duration (sec)']
        s_rows = [[s['screen_name'], s['path'], s['visit_count'], s['unique_users'], s['total_active_mins'], s['avg_duration_sec']] for s in get_screen_analytics()]
        _style_excel_worksheet(ws_s, "SATARK AUDIT - SCREEN USAGE & NAV ANALYTICS", s_headers, s_rows, now_str)

        # 4. Sheet 4: API Diagnostics
        ws_a = wb.create_sheet(title="API Diagnostics")
        a_headers = ['Endpoint Route', 'HTTP Method', 'Total Calls', 'Avg Latency (ms)', 'Error Count', 'Success Rate %']
        a_rows = [[a['endpoint'], a['method'], a['total_calls'], a['avg_latency_ms'], a['error_count'], a['success_rate']] for a in get_api_analytics()]
        _style_excel_worksheet(ws_a, "SATARK AUDIT - API FREQUENCY & PERFORMANCE TELEMETRY", a_headers, a_rows, now_str)

        # 5. Sheet 5: Session Audit Stream
        ws_log = wb.create_sheet(title="Session Audit Logs")
        log_headers = ['Session ID', 'User ID', 'UserCode', 'Login Time', 'Logout Time', 'Active Mins', 'Idle Mins', 'IP Address', 'Browser', 'Logout Reason']
        log_rows = [[log['session_id'], log['user_id'], log['usercode'], log['login_time'], log['logout_time'], log['active_duration_mins'], log['idle_duration_mins'], log['ip'], log['browser'], log['logout_reason']] for log in get_session_logs(limit=500)]
        _style_excel_worksheet(ws_log, "SATARK AUDIT - USER SESSION AUDIT LOG STREAM", log_headers, log_rows, now_str)

        filename = "satark_master_telemetry_report.xlsx"

    elif report_type == 'users':
        ws = wb.active
        ws.title = "User Telemetry"
        u_headers = ['User ID', 'UserCode', 'Total Sessions', 'Total Active Mins', 'Last Login', 'Last Active', 'Top Visited Screen']
        u_rows = [[u['user_id'], u['usercode'], u['total_sessions'], u['total_active_mins'], u['last_login'], u['last_active'], u['top_screen']] for u in get_user_analytics_list()]
        _style_excel_worksheet(ws, "SATARK AUDIT - USER TELEMETRY & ACTIVE TIME REPORT", u_headers, u_rows, now_str)
        filename = "satark_user_telemetry_report.xlsx"

    elif report_type == 'screens':
        ws = wb.active
        ws.title = "Screen Usage"
        s_headers = ['Screen Name', 'Route Path', 'Visit Count', 'Unique Users', 'Total Active Mins', 'Avg Duration (sec)']
        s_rows = [[s['screen_name'], s['path'], s['visit_count'], s['unique_users'], s['total_active_mins'], s['avg_duration_sec']] for s in get_screen_analytics()]
        _style_excel_worksheet(ws, "SATARK AUDIT - SCREEN USAGE ANALYTICS REPORT", s_headers, s_rows, now_str)
        filename = "satark_screen_usage_report.xlsx"

    elif report_type == 'apis':
        ws = wb.active
        ws.title = "API Diagnostics"
        a_headers = ['Endpoint Route', 'HTTP Method', 'Total Calls', 'Avg Latency (ms)', 'Error Count', 'Success Rate %']
        a_rows = [[a['endpoint'], a['method'], a['total_calls'], a['avg_latency_ms'], a['error_count'], a['success_rate']] for a in get_api_analytics()]
        _style_excel_worksheet(ws, "SATARK AUDIT - API FREQUENCY & PERFORMANCE REPORT", a_headers, a_rows, now_str)
        filename = "satark_api_frequency_report.xlsx"

    else:
        ws = wb.active
        ws.title = "Session Audit Logs"
        log_headers = ['Session ID', 'User ID', 'UserCode', 'Login Time', 'Logout Time', 'Active Mins', 'Idle Mins', 'IP Address', 'Browser', 'Logout Reason']
        log_rows = [[log['session_id'], log['user_id'], log['usercode'], log['login_time'], log['logout_time'], log['active_duration_mins'], log['idle_duration_mins'], log['ip'], log['browser'], log['logout_reason']] for log in get_session_logs(limit=500)]
        _style_excel_worksheet(ws, "SATARK AUDIT - USER SESSION AUDIT STREAM REPORT", log_headers, log_rows, now_str)
        filename = "satark_session_audit_report.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


